from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from database import (
    clear_error_logs,
    delete_cells_for_file,
    delete_file_record,
    fetch_file_record,
    get_connection,
    initialize_database,
    insert_cells_batch,
    list_indexed_files,
    log_index_error,
    upsert_file_record,
)
from utils import column_letter, format_modified_at, iter_excel_files, normalize_text, safe_stringify

ProgressCallback = Callable[[int, int, str], None]


@dataclass
class IndexingStats:
    files_found: int = 0
    files_indexed: int = 0
    files_skipped: int = 0
    files_deleted: int = 0
    cells_added: int = 0
    errors: int = 0

    def to_dict(self) -> dict[str, int]:
        return asdict(self)


class ExcelIndexer:
    def __init__(self, db_path: str | Path | None = None) -> None:
        self.db_path = Path(db_path) if db_path else None

    def index_folder(
        self,
        root_folder: str | Path,
        *,
        cleanup_deleted: bool = False,
        progress_callback: ProgressCallback | None = None,
    ) -> IndexingStats:
        root_path = Path(root_folder).expanduser().resolve(strict=False)
        if not root_path.exists() or not root_path.is_dir():
            raise ValueError(f"Папку не знайдено: {root_path}")

        excel_files = iter_excel_files(root_path)
        stats = IndexingStats(files_found=len(excel_files))

        connection = get_connection(self.db_path)
        initialize_database(connection)
        clear_error_logs(connection)

        for index, file_path in enumerate(excel_files, start=1):
            if progress_callback:
                progress_callback(index - 1, len(excel_files), f"Обробка: {file_path.name}")

            try:
                changed = self._index_single_file(connection, file_path)
                if changed:
                    stats.files_indexed += 1
                    stats.cells_added += changed
                else:
                    stats.files_skipped += 1
            except Exception as exc:  # noqa: BLE001
                stats.errors += 1
                log_index_error(
                    connection,
                    file_path=str(file_path),
                    error_message=str(exc),
                    logged_at=datetime.now().astimezone().isoformat(timespec="seconds"),
                )

            if progress_callback:
                progress_callback(index, len(excel_files), f"Оброблено файлів: {index}")

        if cleanup_deleted:
            stats.files_deleted = self.cleanup_deleted_records(
                root_path, connection=connection
            )

        connection.close()
        return stats

    def cleanup_deleted_records(
        self,
        root_folder: str | Path | None = None,
        *,
        connection=None,
    ) -> int:
        owns_connection = connection is None
        connection = connection or get_connection(self.db_path)
        initialize_database(connection)

        root_path = None
        root_prefix = None
        if root_folder:
            root_path = Path(root_folder).expanduser().resolve(strict=False)
            root_prefix = f"{root_path}/"

        deleted = 0
        for row in list_indexed_files(connection, root_prefix=root_prefix):
            file_path = Path(str(row["file_path"])).expanduser().resolve(strict=False)
            if not file_path.exists():
                with connection:
                    delete_file_record(connection, int(row["id"]))
                deleted += 1

        if owns_connection:
            connection.close()
        return deleted

    def _index_single_file(self, connection, file_path: Path) -> int:
        file_path_str = str(file_path)
        modified_at = format_modified_at(file_path)
        existing = fetch_file_record(connection, file_path_str)

        if existing and existing["modified_at"] == modified_at:
            return 0

        total_inserted = 0

        with connection:
            file_id = upsert_file_record(
                connection,
                file_name=file_path.name,
                file_path=file_path_str,
                folder_path=str(file_path.parent),
                modified_at=modified_at,
            )
            delete_cells_for_file(connection, file_id)

            workbook = load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )

            try:
                for worksheet in workbook.worksheets:
                    batch: list[tuple[object, ...]] = []

                    for row in worksheet.iter_rows():
                        for cell in row:
                            value = cell.value
                            if value is None:
                                continue

                            text_value = safe_stringify(value)
                            if not text_value.strip():
                                continue

                            normalized_value = normalize_text(text_value)
                            if not normalized_value:
                                continue

                            batch.append(
                                (
                                    file_id,
                                    worksheet.title,
                                    cell.row,
                                    column_letter(cell.column),
                                    cell.coordinate,
                                    text_value,
                                    normalized_value,
                                )
                            )

                            if len(batch) >= 1000:
                                insert_cells_batch(connection, batch)
                                total_inserted += len(batch)
                                batch.clear()

                    if batch:
                        insert_cells_batch(connection, batch)
                        total_inserted += len(batch)
            finally:
                workbook.close()

        return total_inserted
