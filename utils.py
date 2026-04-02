from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
import re
import subprocess
import unicodedata
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
WHITESPACE_RE = re.compile(r"\s+")


def normalize_text(value: str) -> str:
    """Normalize text for case-insensitive and whitespace-tolerant search."""
    normalized = unicodedata.normalize("NFKC", value)
    normalized = WHITESPACE_RE.sub(" ", normalized)
    return normalized.strip().lower()


def safe_stringify(value: Any) -> str:
    """Convert Excel cell values to a stable string representation."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    return str(value)


def format_modified_at(file_path: Path) -> str:
    """Return file modification time as an ISO string in local time."""
    stat = file_path.stat()
    modified = datetime.fromtimestamp(stat.st_mtime).astimezone()
    return modified.isoformat(timespec="seconds")


def is_excel_file(file_path: Path) -> bool:
    return file_path.is_file() and file_path.suffix.lower() in EXCEL_EXTENSIONS


def is_temporary_excel_file(file_path: Path) -> bool:
    return file_path.name.startswith("~$")


def iter_excel_files(root_folder: Path) -> list[Path]:
    files = [
        path
        for path in root_folder.rglob("*")
        if is_excel_file(path) and not is_temporary_excel_file(path)
    ]
    return sorted(files, key=lambda item: str(item).lower())


def column_letter(column_index: int) -> str:
    return get_column_letter(column_index)


def path_matches_root(file_path: str, root_folder: Path | None) -> bool:
    if root_folder is None:
        return True
    try:
        return Path(file_path).resolve().is_relative_to(root_folder.resolve())
    except FileNotFoundError:
        return Path(file_path).expanduser().resolve(strict=False).is_relative_to(
            root_folder.resolve()
        )


def open_path_in_finder(target: str | Path) -> None:
    path = Path(target).expanduser().resolve(strict=False)
    subprocess.run(["open", str(path)], check=False)


def reveal_file_in_finder(target: str | Path) -> None:
    path = Path(target).expanduser().resolve(strict=False)
    subprocess.run(["open", "-R", str(path)], check=False)


def export_rows_to_excel(
    rows: Sequence[Mapping[str, Any]], sheet_name: str = "Результати"
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or "Результаты"

    if not rows:
        worksheet["A1"] = "Немає даних для експорту"
    else:
        headers = list(rows[0].keys())
        worksheet.append(headers)

        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])

        for index, header in enumerate(headers, start=1):
            max_length = max(
                len(str(header)),
                *[len(str(row.get(header, ""))) for row in rows],
            )
            worksheet.column_dimensions[column_letter(index)].width = min(max_length + 2, 60)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
